from django.shortcuts import render, redirect
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Alignment, Font
import datetime as dt

from django.core.mail import send_mail
from django.conf import settings

from .models import Einkaufsliste

# Create your views here.
def index(request):
    if request.method == 'POST' and 'btn-mail' in request.POST:
        mail_adress = request.POST.get('input-mail', None)

        # db query
        qs = Einkaufsliste.objects.filter(erledigt=False)

        str_einkauf = "\n"
        if qs.count() == 0:
            str_einkauf += "keine Objekte vorhanden"
        else:
            for item in qs:
                str_einkauf += f"- {item.anzahl}x {item.beschreibung} \n"

        send_mail(
            'Einkaufsliste',
            str_einkauf,
            settings.EMAIL_HOST,
            [mail_adress],
            fail_silently=False,
        )
        return redirect('app-grocery-store-index')
    return render(request, 'app_grocery_store/index.html')


def export_xls(request):
    # load data
    qs = Einkaufsliste.objects.filter(erledigt=False)

    # create workbook -> -sheet
    wb = Workbook()

    ws = wb.active
    ws.title = "Einkaufsliste"

    # write data into cells
    ws['A1'] = "Einkaufsliste"
    ws['A2'] = f"Datum: {dt.datetime.now().strftime('%d.%m.%Y')}"

    ws['A4'] = '#'
    ws['B4'] = 'Beschreibung'
    ws['C4'] = 'Anzahl'

    # set text font-size
    ws['A1'].font = Font(size="16", bold=True)
    ws['A2'].font = Font(size="14", bold=True)
    ws['A4'].font = Font(bold=True)
    ws['B4'].font = Font(bold=True)
    ws['C4'].font = Font(bold=True)

    # set column width
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10

    # set alignments inside cell
    ws['A4'].alignment = Alignment(horizontal='center')
    ws['B4'].alignment = Alignment(horizontal='left')
    ws['C4'].alignment = Alignment(horizontal='right')

    for counter, item in enumerate(qs):
        ws[f'A{counter + 5}'] = counter + 1
        ws[f'B{counter + 5}'] = item.beschreibung
        ws[f'C{counter + 5}'] = item.anzahl

        ws[f'A{counter + 5}'].alignment = Alignment(horizontal='center')
        ws[f'B{counter + 5}'].alignment = Alignment(horizontal='left')
        ws[f'C{counter + 5}'].alignment = Alignment(horizontal='right')

    response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=Einkaufsliste - {dt.datetime.now().strftime("%Y%m%d%H%M")}.xlsx'
    return response